# This file is part of [accountancy-support-tools]. 
# # [accountancy-support-tools] is free software: you can redistribute it and/or 
# # modify it under the terms of the GNU General Public License as published by 
# the Free Software Foundation, either version 3 of the License, or (at your option) 
# any later version. 
# # [accountancy-support-tools] is distributed in the hope that it will be useful, 
# but WITHOUT ANY WARRANTY; without even the implied warranty of 
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the 
# GNU General Public License for more details. 
# # You should have received a copy of the GNU General Public License 
# # along with [accountancy-support-tools]. If not, see <http://www.gnu.org/licenses/>.

###############################################################################################

from openpyxl import load_workbook
import os
import datetime
import random
import shutil
from functools import wraps
import sys
import pythoncom
import win32com.client
import win32com.client.gencache
from functools import wraps
import comtypes.client

###############################################################################################

def suppress_errors(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            pass
    return wrapper



def clear_com_cache():
    folder = win32com.client.gencache.GetGeneratePath()
    if os.path.exists(folder):
        shutil.rmtree(folder)



@suppress_errors
def replace_short_text(find, key, value):            
    find.Text = key
    find.Replacement.Text = str(value)
    #------------
    find.Wrap = 1
    find.Forward = True
    find.Format = False
    find.MatchCase = False
    find.MatchWholeWord = False
    find.MatchWildcards = False
    find.MatchSoundsLike = False
    find.MatchAllWordForms = False
    #------------
    find.Execute(Replace=2)
    


@suppress_errors
def replace_long_text(find, key, value):
    max_length = 255
    remaining_value = value

    find.Text = key    
    find.Replacement.Text = remaining_value[:max_length]
    #------------
    find.Wrap = 1
    find.Forward = True
    find.Format = False
    find.MatchCase = False
    find.MatchWholeWord = False
    find.MatchWildcards = False
    find.MatchSoundsLike = False
    find.MatchAllWordForms = False
    #------------  

    if find.Execute(Replace=1):
        remaining_value = remaining_value[max_length:]
        range = find.Parent

        while len(remaining_value) > 0:
            value_chunk = remaining_value[:max_length]
            range.InsertAfter(value_chunk)
            remaining_value = remaining_value[max_length:]
            range.MoveEnd(12, len(value_chunk))        
    else:
        pass



@suppress_errors
def replace_text_in_paragraphs(document_object, key, value):       
    find = document_object.Content.Find
    if len(value) > 255:
        replace_long_text(find, key, value)
    else:
        replace_short_text(find, key, value)



@suppress_errors
def replace_text_in_shapes(document_object, key, value):
    for shape in document_object.Shapes:
        if shape.TextFrame.HasText:
            text_range = shape.TextFrame.TextRange
            find = text_range.Find
            if len(value) > 255:
                replace_long_text(find, key, value)
            else:
                replace_short_text(find, key, value)



@suppress_errors
def replace_text_in_drawings(document_object, key, value):
    for shape in document_object.InlineShapes:
        if shape.Type == 3:
            if shape.TextFrame.HasText:
                text_range = shape.TextFrame.TextRange
                find = text_range.Find
                if len(value) > 255:
                    replace_long_text(find, key, value)
                else:
                    replace_short_text(find, key, value)



@suppress_errors
def replace_text_in_headers_footers(document_object, key, value):
    for section in document_object.Sections:
        for header in section.Headers:
            for paragraph in header.Range.Paragraphs:
                text_range = paragraph.Range
                find = text_range.Find
                if len(value) > 255:
                    replace_long_text(find, key, value)
                else:
                    replace_short_text(find, key, value)

        for footer in section.Footers:
            for paragraph in footer.Range.Paragraphs:
                text_range = paragraph.Range
                find = text_range.Find              
                if len(value) > 255:
                    replace_long_text(find, key, value)
                else:
                    replace_short_text(find, key, value)



@suppress_errors
def replace_text_in_comments(document_object, key, value):
    for comment in document_object.Comments:
        text_range = comment.Range
        find = text_range.Find
        if len(value) > 255:
            replace_long_text(find, key, value)
        else:
            replace_short_text(find, key, value) 



@suppress_errors
def replace_text_in_footnotes_endnotes(document_object, key, value):
    for footnote in document_object.Footnotes:
        text_range = footnote.Range
        find = text_range.Find       
        if len(value) > 255:
            replace_long_text(find, key, value)
        else:
            replace_short_text(find, key, value)

    for endnote in document_object.Endnotes:
        text_range = endnote.Range
        find = text_range.Find
        if len(value) > 255:
            replace_long_text(find, key, value)
        else:
            replace_short_text(find, key, value)



@suppress_errors
def replace_placeholders_with_pywin32(template_path, output_path, data):
    pythoncom.CoInitialize() 
    try:
        word_com_object = comtypes.client.CreateObject("Word.Application")  
        word_com_object.Visible = True

        document_object = word_com_object.Documents.Open(os.path.abspath(template_path))

        for key, value in data.items():
            value = str(value)

            replace_text_in_paragraphs(document_object, key, value)
            replace_text_in_shapes(document_object, key, value)
            replace_text_in_drawings(document_object, key, value)
            replace_text_in_headers_footers(document_object, key, value)
            replace_text_in_comments(document_object, key, value)
            replace_text_in_footnotes_endnotes(document_object, key, value)

        document_object.SaveAs(os.path.abspath(output_path))

    except Exception as e:
        raise e        
    finally:
        if 'document_object' in locals():
            document_object.Close(SaveChanges=False)
        if 'word_com_object' in locals():
            word_com_object.Quit()
            del word_com_object  
        pythoncom.CoUninitialize() 



def generate_word_reports(table_path, template_path, sheet_name):      
    
    filename = os.path.join(f"{table_path}")
    wb = load_workbook(filename, data_only=True)
    ws = wb[sheet_name]
    data = {}
    
    for row in range(3, ws.max_row + 1):
        for col in range(1,ws.max_column + 1):
            key = ws.cell(2, col).value
            value = ws.cell(row, col).value
            if key is not None:
                data[key] = value
            else:
                break
        
        current_datetime = datetime.datetime.now()
        formatted_datetime = current_datetime.strftime("%Y-%m-%d__%H-%M-%S")

        if '{FileName}' in data: 
            output_path = f"output/{data['{FileName}']}__{formatted_datetime}.docx"
        else:
            random_number = str(random.randint(111111, 999999))
            output_path = f"output/{random_number}__{formatted_datetime}.docx"

        replace_placeholders_with_pywin32(template_path, output_path, data)



def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

###############################################################################################

if __name__ == '__main__':
    template_xlsx_path = resource_path('templates/table.xlsx')
    template_docx_path = resource_path('templates/template.docx')
    readme_path = resource_path('templates/README.md')

    table_path = 'input/table.xlsx'
    sheet_name = 'machine_sheet'
    template_path = 'input/template.docx'

    os.makedirs('input', exist_ok=True)
    os.makedirs('output', exist_ok=True)

    if not os.path.exists('input/table.xlsx'):
        shutil.copy(template_xlsx_path, 'input/table.xlsx')
    if not os.path.exists('input/template.docx'):
        shutil.copy(template_docx_path, 'input/template.docx')
    if not os.path.exists('input/README.md'):
        shutil.copy(readme_path, 'input/README.md')

###############################################################################################
   
    clear_com_cache()

    generate_word_reports(table_path, template_path, sheet_name)

###############################################################################################
